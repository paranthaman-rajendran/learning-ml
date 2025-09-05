import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule

def create_dashboard_template():
    """
    Creates an Excel workbook to track application modernization initiatives.
    The template includes predefined columns, sample data, data validation,
    and conditional formatting for status tracking.
    """
    # --- 1. SETUP WORKBOOK AND WORKSHEET ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modernization Dashboard"

    # --- 2. DEFINE HEADERS ---
    headers = [
        "Application Name", "Owner/Team", "Toolset/Tech Being Adopted", "Status",
        "Key Milestones/Dates", "Challenges/Blockers", "Benefits Achieved",
        "Last Updated", "Comments/Notes"
    ]
    ws.append(headers)

    # --- 3. ADD SAMPLE DATA FOR DEMONSTRATION ---
    sample_data = [
        ["CRM Pro", "Sales Ops", "Salesforce Lightning", "Completed", "Q1 2025: Full migration", "", "$15k annual savings", "2025-03-28", "Migration was smooth."],
        ["Inventory Master", "Warehouse Team", "Cloud-based ERP", "In Progress", "Q3 2025: UAT Sign-off", "", "Real-time tracking", "2025-09-01", "Integration testing underway."],
        ["Legacy HR Portal", "HR Department", "Workday", "At Risk", "Q4 2025: Go-live", "Data migration tool is failing on large datasets.", "", "2025-08-30", "Vendor support engaged for blocker."],
        ["Analytics Engine", "BI Team", "Python + Databricks", "In Progress", "Q3 2025: First reports live", "", "Faster data processing", "2025-09-04", ""],
        ["Finance Ledger", "Finance Team", "SAP S/4HANA", "Not Started", "Q1 2026: Project Kick-off", "", "Improved compliance", "", "Awaiting budget approval."]
    ]

    for row in sample_data:
        ws.append(row)

    # --- 4. CREATE AN EXCEL TABLE FOR SORTING/FILTERING ---
    # Define the full range of the data
    table_range = f"A1:I{len(sample_data) + 1}"
    tab = Table(displayName="AppModernizationTable", ref=table_range)

    # Add a default style with striped rows and filtering
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # --- 5. SETUP DATA VALIDATION (DROPDOWN LIST) FOR STATUS COLUMN ---
    # Formula specifies the list items for the dropdown
    dv = DataValidation(type="list", formula1='"Not Started,In Progress,At Risk,Completed"', allow_blank=True)
    # Apply the validation to the entire 'Status' column (D) from row 2 downwards
    dv.add('D2:D1048576')
    ws.add_validation(dv)
    dv.error = 'Your entry is not in the list.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'Status Selection'

    # --- 6. SETUP CONDITIONAL FORMATTING ---
    # Define fill colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Define font color for critical issues
    red_font = Font(color="9C0006")
    critical_issue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


    # Apply rules to the 'Status' column (D)
    status_range = "D2:D1048576"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Completed"'], fill=green_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"At Risk"'], fill=red_fill))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Not Started"'], fill=gray_fill))
    
    # Apply rule to highlight critical issues in 'Challenges/Blockers' column (F)
    # This rule highlights any cell in the column that is NOT empty
    challenges_range = "F2:F1048576"
    ws.conditional_formatting.add(challenges_range, FormulaRule(formula=[f'NOT(ISBLANK({challenges_range.split(":")[0]}))'], stopIfTrue=True, font=red_font, fill=critical_issue_fill))


    # --- 7. ADJUST COLUMN WIDTHS FOR READABILITY ---
    column_widths = {'A': 25, 'B': 20, 'C': 30, 'D': 15, 'E': 30, 'F': 50, 'G': 25, 'H': 15, 'I': 50}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        
    # --- 8. ADD AN INSTRUCTIONS SHEET ---
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80

    instructions_text = [
        ("How to use this Dashboard",),
        ("",),
        ("1. Shared Location:", "Store this file in a shared cloud location like SharePoint, Google Drive, or OneDrive so the whole team can access it."),
        ("2. Regular Updates:", "Owners should update the status of their applications weekly or as significant progress is made."),
        ("3. Status Column:", "Use the dropdown in the 'Status' column to select a standardized status. The cell will automatically change color."),
        ("   - Gray (Not Started): The initiative has not yet begun.",),
        ("   - Yellow (In Progress): Work is underway as planned.",),
        ("   - Red (At Risk): There is a significant blocker or challenge that may impact the timeline or success."),
        ("   - Green (Completed): The transition is complete and benefits are being realized.",),
        ("4. Challenges/Blockers:", "Clearly describe any issues in this column. Any entry will be automatically highlighted in red to draw attention."),
        ("5. Filtering:", "Use the filter arrows in the header row to create executive summaries (e.g., filter to show only 'At Risk' items)."),
    ]

    for row_data in instructions_text:
        ws_instructions.append(row_data)

    # Style the header on the instructions sheet
    ws_instructions['A1'].font = Font(bold=True, size=16)


    # --- 9. SAVE THE WORKBOOK ---
    file_path = "Application_Modernization_Dashboard.xlsx"
    wb.save(file_path)
    print(f"Successfully created '{file_path}'")

if __name__ == "__main__":
    create_dashboard_template()
